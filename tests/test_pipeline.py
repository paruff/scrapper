"""Tests for MultiSheetExcelPipeline writing Excel outputs.

Uses a temporary working directory to avoid touching the repo.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from vrm_crawl.pipelines import MultiSheetExcelPipeline


class DummyLogger:
    def info(self, *args, **kwargs):
        pass

    def warning(self, *args, **kwargs):
        pass


class DummySpider:
    name = "vrm"
    logger = DummyLogger()


def test_excel_pipeline_creates_workbook_and_sheets(tmp_path, monkeypatch):
    # Run in temporary directory so output/ is created there
    monkeypatch.chdir(tmp_path)

    pipe = MultiSheetExcelPipeline()
    spider = DummySpider()
    pipe.open_spider(spider)

    item_va = {
        "state": "VA",
        "name": "A",
        "city": "C",
        "address": "Addr",
        "price": "$1",
        "bedrooms": "1",
        "bathrooms": "1",
        "url": "/a",
    }
    item_tx = {
        "state": "TX",
        "name": "B",
        "city": "Austin",
        "address": "Addr2",
        "price": "$2",
        "bedrooms": "2",
        "bathrooms": "2",
        "url": "/b",
    }

    pipe.process_item(item_va, spider)
    pipe.process_item(item_tx, spider)
    pipe.close_spider(spider)

    date_str = datetime.now().strftime("%Y-%m-%d")
    out = Path("output") / f"vrm_listings_{date_str}.xlsx"
    assert out.exists()

    wb = load_workbook(str(out))
    assert set(["VA", "TX"]).issubset(set(wb.sheetnames))

    # Header row equals item keys order
    va = wb["VA"]
    headers = [c.value for c in next(va.iter_rows(min_row=1, max_row=1))[0:len(item_va)]]
    assert headers == list(item_va.keys())

    # Data rows appended
    assert va.max_row >= 2
