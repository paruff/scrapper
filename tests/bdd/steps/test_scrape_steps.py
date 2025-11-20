from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import load_workbook
from pytest_bdd import given, parsers, scenario, then, when
from scrapy.http import Request, TextResponse

from vrm_crawl.pipelines import MultiSheetExcelPipeline
from vrm_crawl.spiders.vrm import VrmSpider, extract_inline_model


def _build_response(url: str, body: str, meta: dict[str, Any] | None = None) -> TextResponse:
    request = Request(url=url, meta=meta or {})
    return TextResponse(url=url, request=request, body=body.encode("utf-8"), encoding="utf-8")


class DummySettings:
    def __init__(self, values: dict[str, int] | None = None):
        self._values = values or {}

    def getint(self, key: str, default: int) -> int:
        return int(self._values.get(key, default))


@scenario("../features/scrape.feature", "Extract inline JSON model into properties")
def test_extract_inline_model_scenario():
    pass


@scenario("../features/scrape.feature", "Respect per-state page cap and avoid extra follows")
def test_page_cap_no_follow_scenario():
    pass


@scenario("../features/scrape.feature", "Write Excel with per-state sheet and ordered headers")
def test_excel_headers_scenario():
    pass


@given("sample VRM HTML with two properties", target_fixture="sample_html_two_properties")
def sample_html_two_properties() -> str:
    return (
        "<html><head><script>window.__INITIAL_STATE__ = {\n"
        '  "properties": [\n'
        '    {"name":"Beach House Paradise","city":"Virginia Beach","address":"123 Ocean","price":"$2500/week","bedrooms":"4","bathrooms":"3","url":"/beach"},\n'
        '    {"name":"Mountain Retreat","city":"Charlottesville","address":"456 Hill","price":"$1800/week","bedrooms":"3","bathrooms":"2","url":"/mountain"}\n'
        "  ]\n};</script></head><body></body></html>"
    )


@when("I extract the inline model")
def extract_model(sample_html_two_properties, context: dict):  # type: ignore[no-redef]
    context["model"] = extract_inline_model(sample_html_two_properties)


@then("the result has 2 properties")
def assert_two_properties(context: dict):
    model = context.get("model")
    assert model is not None
    assert len(model.get("properties", [])) == 2


@then(parsers.parse('the first property has name "{name}" and city "{city}"'))
def assert_first_property_fields(context: dict, name: str, city: str):
    model = context.get("model")
    first = model["properties"][0]
    assert first["name"] == name
    assert first["city"] == city


@given(parsers.parse('a spider configured with max pages {max_pages:d} for state "{state}"'), target_fixture="configured_spider")
def configured_spider(max_pages: int, state: str) -> VrmSpider:
    spider = VrmSpider(states=state)
    spider.crawler = SimpleNamespace(settings=DummySettings({"VRM_MAX_PAGES": max_pages}))
    list(spider.start_requests())
    return spider


@when("I parse a page containing a next link for that state")
def parse_page_with_next(configured_spider: VrmSpider, context: dict):
    html = (
        "<html><head><script>window.__INITIAL_STATE__ = {\n"
        '  "properties": [{"name":"Only","city":"C","address":"A","price":"$","bedrooms":"1","bathrooms":"1","url":"/only"}]\n'
        "};</script></head>"
        '<body><a class="next-page" href="/va?page=2">Next</a></body></html>'
    )
    resp = _build_response(
        f"https://www.vacationrentalmanagers.com/{configured_spider.states[0].lower()}",
        html,
        meta={"state": configured_spider.states[0]},
    )
    context["parse_out"] = list(configured_spider.parse(resp))


@then("no follow request is produced")
def assert_no_follow(context: dict):
    out = context.get("parse_out", [])
    follows = [o for o in out if isinstance(o, Request)]
    assert follows == []


@given("an open Excel pipeline in a temporary directory", target_fixture="open_pipeline")
def open_pipeline(tmp_path, monkeypatch) -> MultiSheetExcelPipeline:
    monkeypatch.chdir(tmp_path)
    pipe = MultiSheetExcelPipeline()
    pipe.open_spider(type("S", (), {"logger": type("L", (), {"info": lambda *a, **k: None})()})())
    return pipe


@when(parsers.parse('I process an item for state "{state}" with fields in a specific order'))
def process_item_in_order(open_pipeline: MultiSheetExcelPipeline, state: str, context: dict):
    item = {
        "state": state,
        "name": "A",
        "city": "C",
        "address": "Addr",
        "price": "$1",
        "bedrooms": "1",
        "bathrooms": "1",
        "url": "/a",
    }
    context["item_keys"] = list(item.keys())
    open_pipeline.process_item(item, None)
    dummy_spider = type("S", (), {"logger": type("L", (), {"info": lambda *a, **k: None})()})()
    open_pipeline.close_spider(dummy_spider)
    context["output_file"] = Path("output") / f"vrm_listings_{datetime.now().strftime('%Y-%m-%d')}.xlsx"


@then(parsers.parse('the workbook has a sheet "{state}" with headers matching the item key order'))
def assert_headers(context: dict, state: str):
    out = context["output_file"]
    assert out.exists()
    wb = load_workbook(str(out))
    assert state in wb.sheetnames
    ws = wb[state]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == context["item_keys"]
