"""Scrapy Item Pipeline definitions.

MultiSheetExcelPipeline writes property listings to an Excel workbook
with one sheet per state.
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


class MultiSheetExcelPipeline:
    """Pipeline to write scraped items to an Excel workbook.

    Each state gets its own worksheet. The workbook is created if it doesn't
    exist, and the default sheet is removed. Data is appended to sheets.
    """

    def __init__(self):
        self.workbook = None
        self.filename = None
        self.state_rows = {}

    def open_spider(self, spider):
        """Initialize workbook on spider open."""
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        date_str = datetime.now().strftime("%Y-%m-%d")
        self.filename = output_dir / f"vrm_listings_{date_str}.xlsx"

        if self.filename.exists():
            self.workbook = load_workbook(str(self.filename))
        else:
            self.workbook = Workbook()
            # Remove default sheet if it exists
            if "Sheet" in self.workbook.sheetnames:
                del self.workbook["Sheet"]

        spider.logger.info(f"Excel pipeline initialized: {self.filename}")

    def close_spider(self, spider):
        """Save workbook on spider close."""
        if self.workbook:
            self.workbook.save(str(self.filename))
            spider.logger.info(f"Excel workbook saved: {self.filename}")

    def process_item(self, item, spider):
        """Process each item and write to appropriate state sheet."""
        state = item.get("state", "Unknown")

        if state not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(title=state)
            # Write header row
            headers = list(item.keys())
            sheet.append(headers)
            self.state_rows[state] = 1
        else:
            sheet = self.workbook[state]

        # Append data row
        row_data = [item.get(key, "") for key in item.keys()]
        sheet.append(row_data)

        return item
